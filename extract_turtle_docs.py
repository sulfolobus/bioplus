#!/usr/bin/env python3
import os
import sys
import json
from pathlib import Path

def extract_from_pdf(filepath):
    """Extract text from PDF file"""
    try:
        from pypdf import PdfReader
        reader = PdfReader(filepath)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"ERROR: {str(e)}"

def extract_from_docx(filepath):
    """Extract text from Word file"""
    try:
        from docx import Document
        doc = Document(filepath)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text
    except Exception as e:
        return f"ERROR: {str(e)}"

def extract_from_xlsx(filepath):
    """Extract text from Excel file"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except Exception as e:
        return f"ERROR: {str(e)}"

def process_file(filepath):
    """Process a single file based on extension"""
    ext = filepath.suffix.lower()
    
    if ext == '.pdf':
        return extract_from_pdf(filepath)
    elif ext == '.docx':
        return extract_from_docx(filepath)
    elif ext in ['.xlsx', '.xls']:
        return extract_from_xlsx(filepath)
    else:
        return f"Unsupported file type: {ext}"

def main():
    source_dir = Path("/Users/wy/Desktop/龟宠健康知识库")
    output_dir = Path("/Users/wy/Documents/bioplus/turtle-knowledge-base")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    results = []
    
    # Get all supported files
    files = list(source_dir.glob("*.*"))
    files = [f for f in files if f.suffix.lower() in ['.pdf', '.docx', '.xlsx', '.xls']]
    
    print(f"Found {len(files)} files to process")
    
    for i, filepath in enumerate(files, 1):
        print(f"[{i}/{len(files)}] Processing: {filepath.name}")
        
        try:
            content = process_file(filepath)
            
            # Save individual file content
            output_file = output_dir / f"{filepath.stem}.txt"
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(content)
            
            results.append({
                "filename": filepath.name,
                "size": filepath.stat().st_size,
                "output": str(output_file),
                "length": len(content)
            })
            
        except Exception as e:
            print(f"  ERROR: {str(e)}")
            results.append({
                "filename": filepath.name,
                "error": str(e)
            })
    
    # Save summary
    summary_file = output_dir / "summary.json"
    with open(summary_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\nDone! Processed {len(results)} files")
    print(f"Output directory: {output_dir}")

if __name__ == "__main__":
    main()
